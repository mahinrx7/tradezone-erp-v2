import os

from flask import current_app

from werkzeug.utils import secure_filename

from flask_login import login_required

from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    send_file
)

from app.models import (
    Expense,
    Site,
    WorkEntry
)

from app import db

import pandas as pd

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Spacer,
    Paragraph
)

from reportlab.lib import colors

from reportlab.lib.styles import getSampleStyleSheet

from reportlab.lib.pagesizes import A4

from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)

from openpyxl.utils import get_column_letter

from datetime import datetime

import os


expenses = Blueprint(
    "expenses",
    __name__
)


# VIEW EXPENSES
@expenses.route("/expenses")
def view_expenses():

    site_id = request.args.get("site_id")

    start_date = request.args.get("start_date")

    end_date = request.args.get("end_date")

    sites = Site.query.all()

    query = Expense.query


    # SITE FILTER
    if site_id:

        query = query.filter(
            Expense.site_id == site_id
        )


    # DATE FILTER
    if start_date and end_date:

        query = query.filter(
            Expense.date >= start_date,
            Expense.date <= end_date
        )


    expense_list = query.all()


    # EXPENSE TOTAL
    total_expense = 0

    for expense in expense_list:

        total_expense += expense.amount


    # LABOUR COST
    labour_entries = WorkEntry.query.all()

    labour_total = 0

    for entry in labour_entries:

        if site_id:

            if str(entry.site_id) != str(site_id):

                continue

        labour_total += (
            entry.hours *
            entry.labour.hourly_rate
        )


    net_total = total_expense + labour_total


    return render_template(

        "expenses.html",

        expenses=expense_list,

        sites=sites,

        total=round(net_total,2),

        labour_total=round(labour_total,2)
    )


# ADD EXPENSE
@expenses.route("/add_expense", methods=["GET", "POST"])
@login_required
def add_expense():

    sites = Site.query.all()

    if request.method == "POST":

        receipt = request.files.get(
            "receipt"
        )

        filename = None

        if receipt and receipt.filename != "":

            filename = secure_filename(
                receipt.filename
            )

            save_path = os.path.join(

                current_app.config[
                    "UPLOAD_FOLDER"
                ],

                filename
            )

            receipt.save(save_path)


        expense = Expense(

            site_id=request.form["site_id"],

            category=request.form["category"],

            description=request.form["description"],

            amount=float(
                request.form["amount"]
            ),

            date=request.form["date"],

            receipt_image=filename
        )

        db.session.add(expense)

        db.session.commit()

        return redirect("/expenses")


    return render_template(

        "add_expense.html",

        sites=sites
    )

# DELETE EXPENSE
@expenses.route("/delete_expense/<int:id>")
def delete_expense(id):

    expense = Expense.query.get(id)

    db.session.delete(expense)

    db.session.commit()

    return redirect("/expenses")


# EXPORT EXCEL
@expenses.route("/export_expenses_excel")
def export_expenses_excel():

    expenses_list = Expense.query.all()

    data = []

    for expense in expenses_list:

        data.append({

            "Site": expense.site.name,

            "Category": expense.category,

            "Description": expense.description,

            "Amount (BD)": expense.amount,

            "Date": expense.date
        })


    df = pd.DataFrame(data)


    # ABSOLUTE EXPORT PATH
    BASE_DIR = os.path.abspath(
        os.path.dirname(__file__)
    )

    export_folder = os.path.join(
        BASE_DIR,
        "..",
        "exports"
    )

    export_folder = os.path.abspath(
        export_folder
    )


    if not os.path.exists(export_folder):

        os.makedirs(export_folder)


    filename = f"expenses_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    file_path = os.path.join(
        export_folder,
        filename
    )


    with pd.ExcelWriter(
        file_path,
        engine="openpyxl"
    ) as writer:

        df.to_excel(
            writer,
            sheet_name="Expenses",
            index=False
        )

        workbook = writer.book

        worksheet = writer.sheets["Expenses"]


        # HEADER STYLE
        header_fill = PatternFill(

            start_color="1F2937",

            end_color="1F2937",

            fill_type="solid"
        )

        header_font = Font(

            bold=True,

            color="FFFFFF",

            size=12
        )

        thin = Side(

            border_style="thin",

            color="D1D5DB"
        )

        border = Border(

            left=thin,

            right=thin,

            top=thin,

            bottom=thin
        )


        # HEADER
        for cell in worksheet[1]:

            cell.fill = header_fill

            cell.font = header_font

            cell.border = border

            cell.alignment = Alignment(

                horizontal="center",

                vertical="center"
            )


        # BODY
        for row in worksheet.iter_rows(min_row=2):

            for cell in row:

                cell.border = border

                cell.alignment = Alignment(

                    vertical="center"
                )


        # COLUMN WIDTHS
        column_widths = {

            1:20,
            2:20,
            3:35,
            4:18,
            5:18
        }

        for col_num, width in column_widths.items():

            col_letter = get_column_letter(
                col_num
            )

            worksheet.column_dimensions[
                col_letter
            ].width = width


        # ROW HEIGHTS
        for row in worksheet.iter_rows():

            worksheet.row_dimensions[
                row[0].row
            ].height = 24


    return send_file(

        file_path,

        as_attachment=True
    )


# EXPORT PDF
@expenses.route("/export_expenses_pdf")
def export_expenses_pdf():

    expenses_list = Expense.query.all()


    # ABSOLUTE EXPORT PATH
    BASE_DIR = os.path.abspath(
        os.path.dirname(__file__)
    )

    export_folder = os.path.join(
        BASE_DIR,
        "..",
        "exports"
    )

    export_folder = os.path.abspath(
        export_folder
    )


    if not os.path.exists(export_folder):

        os.makedirs(export_folder)


    filename = f"expenses_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    pdf_path = os.path.join(
        export_folder,
        filename
    )


    doc = SimpleDocTemplate(

        pdf_path,

        pagesize=A4
    )

    styles = getSampleStyleSheet()

    elements = []


    title = Paragraph(

        "Tradezone ERP - Expense Report",

        styles["Heading1"]
    )

    elements.append(title)

    elements.append(
        Spacer(1,20)
    )


    data = [[

        "Site",
        "Category",
        "Description",
        "Amount",
        "Date"

    ]]


    for expense in expenses_list:

        data.append([

            expense.site.name,

            expense.category,

            expense.description,

            f"BD {expense.amount}",

            expense.date
        ])


    table = Table(data)


    table.setStyle(TableStyle([

        # HEADER
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F2937")),

        ("TEXTCOLOR",(0,0),(-1,0),colors.white),

        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

        ("FONTSIZE",(0,0),(-1,0),12),

        ("BOTTOMPADDING",(0,0),(-1,0),12),

        # BODY
        ("BACKGROUND",(0,1),(-1,-1),colors.whitesmoke),

        ("TEXTCOLOR",(0,1),(-1,-1),colors.black),

        ("FONTNAME",(0,1),(-1,-1),"Helvetica"),

        ("FONTSIZE",(0,1),(-1,-1),10),

        # GRID
        ("GRID",(0,0),(-1,-1),1,colors.grey),

        # ALIGN
        ("ALIGN",(0,0),(-1,-1),"CENTER"),

        # PADDING
        ("TOPPADDING",(0,0),(-1,-1),8),

        ("BOTTOMPADDING",(0,0),(-1,-1),8)

    ]))


    elements.append(table)

    doc.build(elements)


    return send_file(

        pdf_path,

        as_attachment=True
    )

@expenses.route("/edit_expense/<int:expense_id>", methods=["GET", "POST"])
@login_required
def edit_expense(expense_id):

    expense = Expense.query.get_or_404(expense_id)

    sites = Site.query.all()

    if request.method == "POST":

        expense.site_id = request.form["site_id"]

        expense.category = request.form["category"]

        expense.description = request.form["description"]

        expense.amount = request.form["amount"]

        expense.date = datetime.strptime(
            request.form["date"],
            "%Y-%m-%d"
        )

        db.session.commit()

        flash("Expense updated successfully")

        return redirect(url_for("expenses.view_expenses"))

    return render_template(

        "edit_expense.html",

        expense=expense,

        sites=sites
    )